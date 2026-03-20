"""
python scripts/utils/export_configs_outputs_to_excel.py --configs_root configs --outputs_root outputs --out outputs/experiment_summary.xlsx
"""
from __future__ import annotations

import json
import re
import argparse
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================
# Basic I/O
# ============================================================

def load_json(path: Path) -> Optional[Dict[str, Any]]:
    if not path.exists():
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def load_yaml(path: Path) -> Optional[Dict[str, Any]]:
    if not path.exists():
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f)
    except Exception:
        return None


def load_first_existing(paths: List[Path]) -> Optional[Dict[str, Any]]:
    for p in paths:
        if p.suffix.lower() == ".json":
            obj = load_json(p)
        else:
            obj = load_yaml(p)
        if obj is not None:
            return obj
    return None


# ============================================================
# Generic helpers
# ============================================================

def safe_get(d: Optional[Dict[str, Any]], *keys, default=None):
    cur = d
    for k in keys:
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    return cur


def to_float(x: Any) -> Optional[float]:
    try:
        if x is None:
            return None
        return float(x)
    except Exception:
        return None


def to_int(x: Any) -> Optional[int]:
    try:
        if x is None:
            return None
        return int(x)
    except Exception:
        return None


def join_list(x: Any) -> Optional[str]:
    if x is None:
        return None
    if isinstance(x, (list, tuple)):
        return ",".join(str(v) for v in x)
    return str(x)


def relpath_str(path: Path, base: Path) -> str:
    try:
        return str(path.relative_to(base))
    except Exception:
        return str(path)


def normalize_slashes(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    return str(s).replace("\\", "/").strip()


def normalize_output_dir_string(s: Optional[str], outputs_root_name: str = "outputs") -> Optional[str]:
    """
    Convert strings like:
      outputs/lm/exp1/...
      .\\outputs\\lm\\exp1\\...
      lm/exp1/...
    into:
      lm/exp1/...
    """
    s = normalize_slashes(s)
    if not s:
        return None

    s = re.sub(r"^\./", "", s)
    s = re.sub(r"^[A-Za-z]:/", "", s)  # strip Windows drive if present
    s = re.sub(r"^/+", "", s)

    prefix = outputs_root_name.rstrip("/") + "/"
    if s.startswith(prefix):
        s = s[len(prefix):]

    s = re.sub(r"/+", "/", s).strip("/")
    return s or None


def parse_rank_alpha_from_text(text: str) -> Tuple[Optional[int], Optional[float]]:
    """
    Parse:
      r64_ar1.25 -> rank=64, alpha=80
      r32_ar1    -> rank=32, alpha=32
    """
    if not text:
        return None, None

    m_r = re.search(r"(?:^|[_/\\-])r(\d+)(?:$|[_/\\-])", text)
    m_ar = re.search(r"(?:^|[_/\\-])ar([0-9]*\.?[0-9]+)(?:$|[_/\\-])", text)

    rank = int(m_r.group(1)) if m_r else None
    ar = float(m_ar.group(1)) if m_ar else None

    if rank is not None and ar is not None:
        return rank, rank * ar
    return rank, None


def parse_tm_tag(text: str) -> Optional[str]:
    if not text:
        return None
    m = re.search(r"(tm-[A-Za-z0-9]+)", text)
    return m.group(1) if m else None


def parse_seed_from_text(text: str) -> Optional[int]:
    if not text:
        return None
    m = re.search(r"(?:^|[_/\\-])seed(\d+)(?:$|[_/\\-])", text.lower())
    return int(m.group(1)) if m else None


def infer_task_family_from_parts(parts: List[str], task_type: Optional[str]) -> Optional[str]:
    parts = [p.lower() for p in parts]
    if "cls" in parts:
        return "cls"
    if "lm" in parts:
        return "lm"

    tt = str(task_type or "").lower()
    if "causal_lm" in tt or tt == "lm":
        return "lm"
    if "classification" in tt:
        return "cls"
    return None


def infer_exp_name_from_parts(parts: List[str]) -> Optional[str]:
    for p in parts:
        if re.fullmatch(r"exp\d+", p.lower()):
            return p.lower()
    return None


def infer_method_from_config_and_path(
    path_str: str,
    cfg: Optional[Dict[str, Any]],
    meta: Optional[Dict[str, Any]],
) -> str:
    """
    Prefer config/meta first, then path.
    Important: check eora before lora.
    """
    p = (path_str or "").lower()

    # Strongest signals
    if safe_get(cfg, "eora") is not None:
        if "quant" in p or safe_get(cfg, "quantized_model_dir") is not None or safe_get(meta, "quantized_model_dir") is not None:
            return "eora_quant"
        return "eora"

    if safe_get(cfg, "lora") is not None:
        if safe_get(cfg, "teacher_model_dir") is not None or "kd" in p:
            return "lora_kd"
        if "quant" in p or safe_get(cfg, "quantized_model_dir") is not None:
            return "lora_quant"
        return "lora"

    if safe_get(meta, "baseline_quant_test") is not None:
        return "eora_quant"

    # Path fallback
    if "optimized" in p and "exp0" in p:
        return "optimized"

    if "quant" in p and "eora" in p:
        return "eora_quant"
    if "quant" in p and "lora" in p:
        return "lora_quant"
    if "quant" in p:
        return "quantized"

    if "kd" in p:
        return "lora_kd"

    if "eora" in p:
        return "eora"
    if "lora" in p:
        return "lora"

    return "unknown"


# ============================================================
# Config scanning
# ============================================================

def should_skip_config(path: Path, configs_root: Path, include_archive: bool) -> bool:
    s = str(path).replace("\\", "/").lower()
    if s.endswith(".bak"):
        return True
    if not include_archive and "/archive/" in s:
        return True
    return False


def build_config_row(config_path: Path, configs_root: Path, outputs_root_name: str) -> Dict[str, Any]:
    cfg = load_yaml(config_path) or {}
    rel = relpath_str(config_path, configs_root).replace("\\", "/")
    parts = rel.split("/")

    task_type = safe_get(cfg, "task_type", default=None)
    task_family = infer_task_family_from_parts(parts, task_type)
    exp_name = infer_exp_name_from_parts(parts)

    path_str = rel.lower()
    method = infer_method_from_config_and_path(path_str, cfg, None)

    output_dir = safe_get(cfg, "output_dir", default=None)
    output_dir_norm = normalize_output_dir_string(output_dir, outputs_root_name=outputs_root_name)

    export_best_model_dir = safe_get(cfg, "export_best_model_dir", default=None)
    export_best_model_dir_norm = normalize_output_dir_string(export_best_model_dir, outputs_root_name=outputs_root_name)

    quant_output_dir = safe_get(cfg, "quant_output_dir", default=None)
    quant_output_dir_norm = normalize_output_dir_string(quant_output_dir, outputs_root_name=outputs_root_name)

    rank = to_int(safe_get(cfg, "lora", "rank"))
    if rank is None:
        rank = to_int(safe_get(cfg, "eora", "rank"))

    alpha = to_float(safe_get(cfg, "lora", "alpha"))
    if alpha is None:
        alpha = to_float(safe_get(cfg, "eora", "alpha"))

    if rank is None or alpha is None:
        parsed_rank, parsed_alpha = parse_rank_alpha_from_text(path_str)
        if rank is None:
            rank = parsed_rank
        if alpha is None:
            alpha = parsed_alpha

    alpha_over_rank = (alpha / rank) if (alpha is not None and rank not in (None, 0)) else None

    seed = safe_get(cfg, "seed", default=None)
    if seed is None:
        seed = parse_seed_from_text(path_str)

    target_modules = safe_get(cfg, "lora", "target_modules")
    if target_modules is None:
        target_modules = safe_get(cfg, "eora", "target_modules")
    target_modules = join_list(target_modules)

    target_modules_tag = parse_tm_tag(path_str)

    data_cfg = safe_get(cfg, "data", default={}) or {}
    train_cfg = safe_get(cfg, "train", default={}) or {}
    kd_cfg = safe_get(cfg, "kd", default={}) or {}
    distill_cfg = safe_get(cfg, "distill", default={}) or {}

    row = {
        "record_type": "config",
        "match_status": "config_only",

        "config_rel_path": rel,
        "config_abs_path": str(config_path),
        "config_name": config_path.name,

        "task_family": task_family,
        "exp_name": exp_name,
        "method": method,

        "model_name": safe_get(cfg, "model_name", default=None),
        "task_type": task_type,
        "seed": seed,
        "num_labels": safe_get(cfg, "num_labels", default=None),

        "rank": rank,
        "alpha": alpha,
        "alpha_over_rank": alpha_over_rank,
        "dropout": safe_get(cfg, "lora", "dropout", default=safe_get(cfg, "eora", "dropout", default=None)),

        "target_modules_tag": target_modules_tag,
        "target_modules": target_modules,

        "output_dir": output_dir,
        "output_dir_norm": output_dir_norm,
        "quant_output_dir": quant_output_dir,
        "quant_output_dir_norm": quant_output_dir_norm,
        "export_best_model_dir": export_best_model_dir,
        "export_best_model_dir_norm": export_best_model_dir_norm,

        "optimized_model_dir": safe_get(cfg, "optimized_model_dir", default=None),
        "teacher_model_dir": safe_get(cfg, "teacher_model_dir", default=None),
        "quantized_model_dir": safe_get(cfg, "quantized_model_dir", default=None),

        "batch_size": safe_get(data_cfg, "batch_size", default=None),
        "max_length": safe_get(data_cfg, "max_length", default=None),
        "num_workers": safe_get(data_cfg, "num_workers", default=None),
        "pin_memory": safe_get(data_cfg, "pin_memory", default=None),
        "persistent_workers": safe_get(data_cfg, "persistent_workers", default=None),
        "prefetch_factor": safe_get(data_cfg, "prefetch_factor", default=None),
        "pad_to_multiple_of": safe_get(data_cfg, "pad_to_multiple_of", default=None),

        "dataset_name": safe_get(data_cfg, "dataset_name", default=None),
        "train_split": safe_get(data_cfg, "train_corpus", "split", default=safe_get(data_cfg, "train_split", default=None)),
        "eval_split": safe_get(data_cfg, "eval_corpus", "split", default=safe_get(data_cfg, "eval_split", default=None)),
        "test_split": safe_get(data_cfg, "test_corpus", "split", default=safe_get(data_cfg, "test_split", default=None)),
        "train_local_disk_path": safe_get(data_cfg, "train_corpus", "local_disk_path", default=None),
        "eval_local_disk_path": safe_get(data_cfg, "eval_corpus", "local_disk_path", default=None),
        "test_local_disk_path": safe_get(data_cfg, "test_corpus", "local_disk_path", default=None),
        "train_text_key": safe_get(data_cfg, "train_corpus", "text_key", default=None),
        "eval_text_key": safe_get(data_cfg, "eval_corpus", "text_key", default=None),
        "test_text_key": safe_get(data_cfg, "test_corpus", "text_key", default=None),

        "lr": safe_get(train_cfg, "lr", default=None),
        "weight_decay": safe_get(train_cfg, "weight_decay", default=None),
        "num_epochs": safe_get(train_cfg, "num_epochs", default=None),
        "max_train_steps": safe_get(train_cfg, "max_train_steps", default=None),
        "grad_accum_steps": safe_get(train_cfg, "grad_accum_steps", default=None),
        "warmup_ratio": safe_get(train_cfg, "warmup_ratio", default=None),
        "scheduler": safe_get(train_cfg, "scheduler", default=None),
        "use_amp": safe_get(train_cfg, "use_amp", default=None),
        "grad_clip": safe_get(train_cfg, "grad_clip", default=None),
        "log_every_steps": safe_get(train_cfg, "log_every_steps", default=safe_get(train_cfg, "log_every", default=None)),
        "eval_every_steps": safe_get(train_cfg, "eval_every_steps", default=safe_get(train_cfg, "eval_every", default=None)),

        "kd_T": safe_get(kd_cfg, "T", default=None),
        "kd_lambda": safe_get(kd_cfg, "lambda", default=None),
        "kd_loss": safe_get(kd_cfg, "loss", default=None),

        "distill_temperature": safe_get(distill_cfg, "temperature", default=None),

        "calibration_num_samples": safe_get(cfg, "calibration_num_samples", default=None),

        # placeholders for output-derived fields
        "outputs_rel_dir": None,
        "outputs_abs_dir": None,
        "run_name": None,
        "teacher_alias": None,

        "val_loss": None,
        "val_ppl": None,
        "val_accuracy": None,
        "val_acc": None,
        "val_f1": None,
        "val_macro_f1": None,
        "val_ce_loss": None,
        "val_kl_to_teacher": None,
        "val_mse_logits_to_teacher": None,

        "test_loss": None,
        "test_ppl": None,
        "test_accuracy": None,
        "test_acc": None,
        "test_f1": None,
        "test_macro_f1": None,
        "test_ce_loss": None,
        "test_kl_to_teacher": None,
        "test_mse_logits_to_teacher": None,

        "baseline_quant_test_loss": None,
        "baseline_quant_test_ppl": None,
        "baseline_quant_test_accuracy": None,
        "baseline_quant_test_acc": None,
        "baseline_quant_test_f1": None,

        "metrics_json": None,
        "meta_json": None,
        "run_info_json": None,
        "config_used_json": None,
        "config_used_yaml": None,
    }
    return row


# ============================================================
# Output scanning
# ============================================================

def extract_metric_block(metrics: Optional[Dict[str, Any]], key: str) -> Dict[str, Any]:
    if not isinstance(metrics, dict):
        return {}
    block = metrics.get(key, {})
    return block if isinstance(block, dict) else {}


def pick_test_block(metrics: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not isinstance(metrics, dict):
        return {}
    if isinstance(metrics.get("test"), dict):
        return metrics["test"]
    if isinstance(metrics.get("test.py"), dict):
        return metrics["test.py"]
    return {}


def build_output_row(metrics_path: Path, outputs_root: Path) -> Dict[str, Any]:
    run_dir = metrics_path.parent

    metrics = load_json(metrics_path) or {}
    meta = load_json(run_dir / "meta.json") or {}
    run_info = load_json(run_dir / "run_info.json") or {}
    config = load_first_existing([
        run_dir / "config_used.json",
        run_dir / "config_used.yaml",
        run_dir / "config.yaml",
    ]) or {}

    run_dir_rel = relpath_str(run_dir, outputs_root).replace("\\", "/")
    path_str = run_dir_rel.lower()
    parts = run_dir_rel.split("/")

    task_type = (
        safe_get(metrics, "task_type", default=None)
        or safe_get(meta, "task_type", default=None)
        or safe_get(run_info, "task_type", default=None)
        or safe_get(config, "task_type", default=None)
    )

    task_family = infer_task_family_from_parts(parts, task_type)
    exp_name = infer_exp_name_from_parts(parts)
    method = infer_method_from_config_and_path(path_str, config, meta)

    rank = to_int(safe_get(config, "lora", "rank"))
    if rank is None:
        rank = to_int(safe_get(config, "eora", "rank"))

    alpha = to_float(safe_get(config, "lora", "alpha"))
    if alpha is None:
        alpha = to_float(safe_get(config, "eora", "alpha"))

    if rank is None or alpha is None:
        parsed_rank, parsed_alpha = parse_rank_alpha_from_text(path_str)
        if rank is None:
            rank = parsed_rank
        if alpha is None:
            alpha = parsed_alpha

    alpha_over_rank = (alpha / rank) if (alpha is not None and rank not in (None, 0)) else None

    target_modules = safe_get(config, "lora", "target_modules")
    if target_modules is None:
        target_modules = safe_get(config, "eora", "target_modules")
    target_modules = join_list(target_modules)

    target_modules_tag = parse_tm_tag(path_str)

    seed = safe_get(metrics, "seed", default=None)
    if seed is None:
        seed = safe_get(meta, "seed", default=None)
    if seed is None:
        seed = safe_get(run_info, "seed", default=None)
    if seed is None:
        seed = safe_get(config, "seed", default=None)
    if seed is None:
        seed = parse_seed_from_text(path_str)

    model_name = (
        safe_get(metrics, "model_name", default=None)
        or safe_get(meta, "model_name", default=None)
        or safe_get(run_info, "model_name", default=None)
        or safe_get(config, "model_name", default=None)
    )

    run_name = (
        safe_get(metrics, "run_name", default=None)
        or safe_get(meta, "run_name", default=None)
        or safe_get(run_info, "run_name", default=None)
        or run_dir.name
    )

    teacher_dir = (
        safe_get(meta, "teacher_model_dir", default=None)
        or safe_get(config, "teacher_model_dir", default=None)
        or safe_get(meta, "optimized_model_dir", default=None)
        or safe_get(config, "optimized_model_dir", default=None)
    )
    teacher_alias = None
    if teacher_dir:
        teacher_alias = Path(str(teacher_dir)).name
        parent_name = Path(str(teacher_dir)).parent.name
        if teacher_alias == "model" and parent_name:
            teacher_alias = parent_name

    val_block = extract_metric_block(metrics, "val")
    test_block = pick_test_block(metrics)
    baseline_quant_test = safe_get(meta, "baseline_quant_test", default={}) or {}

    data_cfg = safe_get(config, "data", default={}) or {}
    train_cfg = safe_get(config, "train", default={}) or {}
    kd_cfg = safe_get(config, "kd", default={}) or {}
    distill_cfg = safe_get(config, "distill", default={}) or {}

    # infer base output dir from config if available
    output_dir = safe_get(config, "output_dir", default=None)
    output_dir_norm = normalize_output_dir_string(output_dir, outputs_root_name=outputs_root.name)

    row = {
        "record_type": "output",
        "match_status": "output_only",

        "config_rel_path": None,
        "config_abs_path": None,
        "config_name": None,

        "task_family": task_family,
        "exp_name": exp_name,
        "method": method,

        "model_name": model_name,
        "task_type": task_type,
        "seed": seed,
        "num_labels": safe_get(config, "num_labels", default=safe_get(meta, "num_labels", default=None)),

        "rank": rank,
        "alpha": alpha,
        "alpha_over_rank": alpha_over_rank,
        "dropout": safe_get(config, "lora", "dropout", default=safe_get(config, "eora", "dropout", default=None)),

        "target_modules_tag": target_modules_tag,
        "target_modules": target_modules,

        "output_dir": output_dir,
        "output_dir_norm": output_dir_norm,
        "quant_output_dir": safe_get(config, "quant_output_dir", default=None),
        "quant_output_dir_norm": normalize_output_dir_string(safe_get(config, "quant_output_dir", default=None), outputs_root_name=outputs_root.name),
        "export_best_model_dir": safe_get(config, "export_best_model_dir", default=None),
        "export_best_model_dir_norm": normalize_output_dir_string(safe_get(config, "export_best_model_dir", default=None), outputs_root_name=outputs_root.name),

        "optimized_model_dir": safe_get(config, "optimized_model_dir", default=safe_get(meta, "optimized_model_dir", default=None)),
        "teacher_model_dir": safe_get(config, "teacher_model_dir", default=safe_get(meta, "teacher_model_dir", default=None)),
        "quantized_model_dir": safe_get(config, "quantized_model_dir", default=safe_get(meta, "quantized_model_dir", default=None)),

        "batch_size": safe_get(data_cfg, "batch_size", default=None),
        "max_length": safe_get(data_cfg, "max_length", default=None),
        "num_workers": safe_get(data_cfg, "num_workers", default=None),
        "pin_memory": safe_get(data_cfg, "pin_memory", default=None),
        "persistent_workers": safe_get(data_cfg, "persistent_workers", default=None),
        "prefetch_factor": safe_get(data_cfg, "prefetch_factor", default=None),
        "pad_to_multiple_of": safe_get(data_cfg, "pad_to_multiple_of", default=None),

        "dataset_name": safe_get(data_cfg, "dataset_name", default=None),
        "train_split": safe_get(data_cfg, "train_corpus", "split", default=safe_get(data_cfg, "train_split", default=None)),
        "eval_split": safe_get(data_cfg, "eval_corpus", "split", default=safe_get(data_cfg, "eval_split", default=None)),
        "test_split": safe_get(data_cfg, "test_corpus", "split", default=safe_get(data_cfg, "test_split", default=None)),
        "train_local_disk_path": safe_get(data_cfg, "train_corpus", "local_disk_path", default=None),
        "eval_local_disk_path": safe_get(data_cfg, "eval_corpus", "local_disk_path", default=None),
        "test_local_disk_path": safe_get(data_cfg, "test_corpus", "local_disk_path", default=None),
        "train_text_key": safe_get(data_cfg, "train_corpus", "text_key", default=None),
        "eval_text_key": safe_get(data_cfg, "eval_corpus", "text_key", default=None),
        "test_text_key": safe_get(data_cfg, "test_corpus", "text_key", default=None),

        "lr": safe_get(train_cfg, "lr", default=None),
        "weight_decay": safe_get(train_cfg, "weight_decay", default=None),
        "num_epochs": safe_get(train_cfg, "num_epochs", default=None),
        "max_train_steps": safe_get(train_cfg, "max_train_steps", default=None),
        "grad_accum_steps": safe_get(train_cfg, "grad_accum_steps", default=None),
        "warmup_ratio": safe_get(train_cfg, "warmup_ratio", default=None),
        "scheduler": safe_get(train_cfg, "scheduler", default=None),
        "use_amp": safe_get(train_cfg, "use_amp", default=None),
        "grad_clip": safe_get(train_cfg, "grad_clip", default=None),
        "log_every_steps": safe_get(train_cfg, "log_every_steps", default=safe_get(train_cfg, "log_every", default=None)),
        "eval_every_steps": safe_get(train_cfg, "eval_every_steps", default=safe_get(train_cfg, "eval_every", default=None)),

        "kd_T": safe_get(kd_cfg, "T", default=None),
        "kd_lambda": safe_get(kd_cfg, "lambda", default=None),
        "kd_loss": safe_get(kd_cfg, "loss", default=None),

        "distill_temperature": safe_get(distill_cfg, "temperature", default=None),

        "calibration_num_samples": safe_get(config, "calibration_num_samples", default=None),

        "outputs_rel_dir": run_dir_rel,
        "outputs_abs_dir": str(run_dir),
        "run_name": run_name,
        "teacher_alias": teacher_alias,

        "val_loss": safe_get(val_block, "loss", default=None),
        "val_ppl": safe_get(val_block, "ppl", default=None),
        "val_accuracy": safe_get(val_block, "accuracy", default=None),
        "val_acc": safe_get(val_block, "acc", default=None),
        "val_f1": safe_get(val_block, "f1", default=None),
        "val_macro_f1": safe_get(val_block, "macro_f1", default=None),
        "val_ce_loss": safe_get(val_block, "ce_loss", default=None),
        "val_kl_to_teacher": safe_get(val_block, "kl_to_teacher", default=None),
        "val_mse_logits_to_teacher": safe_get(val_block, "mse_logits_to_teacher", default=None),

        "test_loss": safe_get(test_block, "loss", default=None),
        "test_ppl": safe_get(test_block, "ppl", default=None),
        "test_accuracy": safe_get(test_block, "accuracy", default=None),
        "test_acc": safe_get(test_block, "acc", default=None),
        "test_f1": safe_get(test_block, "f1", default=None),
        "test_macro_f1": safe_get(test_block, "macro_f1", default=None),
        "test_ce_loss": safe_get(test_block, "ce_loss", default=None),
        "test_kl_to_teacher": safe_get(test_block, "kl_to_teacher", default=None),
        "test_mse_logits_to_teacher": safe_get(test_block, "mse_logits_to_teacher", default=None),

        "baseline_quant_test_loss": safe_get(baseline_quant_test, "loss", default=None),
        "baseline_quant_test_ppl": safe_get(baseline_quant_test, "ppl", default=None),
        "baseline_quant_test_accuracy": safe_get(baseline_quant_test, "accuracy", default=None),
        "baseline_quant_test_acc": safe_get(baseline_quant_test, "acc", default=None),
        "baseline_quant_test_f1": safe_get(baseline_quant_test, "f1", default=None),

        "metrics_json": relpath_str(metrics_path, outputs_root),
        "meta_json": relpath_str(run_dir / "meta.json", outputs_root) if (run_dir / "meta.json").exists() else None,
        "run_info_json": relpath_str(run_dir / "run_info.json", outputs_root) if (run_dir / "run_info.json").exists() else None,
        "config_used_json": relpath_str(run_dir / "config_used.json", outputs_root) if (run_dir / "config_used.json").exists() else None,
        "config_used_yaml": relpath_str(run_dir / "config_used.yaml", outputs_root) if (run_dir / "config_used.yaml").exists() else None,
    }
    return row


# ============================================================
# Matching configs <-> outputs
# ============================================================

def is_output_under_config(output_rel_dir: Optional[str], output_dir_norm: Optional[str]) -> bool:
    if not output_rel_dir or not output_dir_norm:
        return False
    a = normalize_slashes(output_rel_dir).strip("/")
    b = normalize_slashes(output_dir_norm).strip("/")
    return a == b or a.startswith(b + "/")


def score_config_output_match(cfg_row: Dict[str, Any], out_row: Dict[str, Any]) -> int:
    score = 0

    if is_output_under_config(out_row.get("outputs_rel_dir"), cfg_row.get("output_dir_norm")):
        score += 100

    if cfg_row.get("task_family") == out_row.get("task_family"):
        score += 10

    if cfg_row.get("exp_name") == out_row.get("exp_name"):
        score += 10

    if cfg_row.get("method") == out_row.get("method"):
        score += 20

    if cfg_row.get("rank") is not None and cfg_row.get("rank") == out_row.get("rank"):
        score += 20

    if cfg_row.get("alpha") is not None and out_row.get("alpha") is not None:
        if abs(float(cfg_row["alpha"]) - float(out_row["alpha"])) < 1e-9:
            score += 20

    if cfg_row.get("seed") is not None and cfg_row.get("seed") == out_row.get("seed"):
        score += 10

    if cfg_row.get("target_modules_tag") and cfg_row.get("target_modules_tag") == out_row.get("target_modules_tag"):
        score += 10

    if cfg_row.get("target_modules") and cfg_row.get("target_modules") == out_row.get("target_modules"):
        score += 10

    return score


def merge_config_output(cfg_row: Dict[str, Any], out_row: Dict[str, Any]) -> Dict[str, Any]:
    """
    Output values take precedence for actual results.
    Config values fill in missing fields.
    """
    merged = dict(cfg_row)
    for k, v in out_row.items():
        if v is not None:
            merged[k] = v

    merged["record_type"] = "merged"
    merged["match_status"] = "matched"
    return merged


def build_merged_records(config_rows: List[Dict[str, Any]], output_rows: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    merged_records: List[Dict[str, Any]] = []
    configs_missing_outputs: List[Dict[str, Any]] = []
    outputs_missing_configs: List[Dict[str, Any]] = []

    used_config_idx = set()
    used_output_idx = set()

    for oi, out_row in enumerate(output_rows):
        best_idx = None
        best_score = -1

        for ci, cfg_row in enumerate(config_rows):
            score = score_config_output_match(cfg_row, out_row)
            if score > best_score:
                best_score = score
                best_idx = ci

        if best_idx is not None and best_score >= 100:
            merged = merge_config_output(config_rows[best_idx], out_row)
            merged_records.append(merged)
            used_config_idx.add(best_idx)
            used_output_idx.add(oi)
        else:
            outputs_missing_configs.append(out_row)

    for ci, cfg_row in enumerate(config_rows):
        if ci not in used_config_idx:
            configs_missing_outputs.append(cfg_row)

    # include unmatched rows into merged_records too, with status
    merged_records.extend(outputs_missing_configs)
    merged_records.extend(configs_missing_outputs)

    return merged_records, configs_missing_outputs, outputs_missing_configs


# ============================================================
# Best-by-group
# ============================================================

def best_metric_value(row: Dict[str, Any]) -> Tuple[Optional[float], bool, str]:
    task_family = row.get("task_family")

    if task_family == "lm":
        val = to_float(row.get("test_ppl"))
        return val, False, "test_ppl"

    for name in ["test_accuracy", "test_acc", "test_f1", "test_macro_f1"]:
        val = to_float(row.get(name))
        if val is not None:
            return val, True, name

    val = to_float(row.get("test_loss"))
    return val, False, "test_loss"


def build_group_key(row: Dict[str, Any]) -> Tuple[Any, ...]:
    return (
        row.get("task_family"),
        row.get("exp_name"),
        row.get("method"),
        row.get("target_modules_tag"),
        row.get("target_modules"),
        row.get("teacher_alias"),
    )


def build_best_by_group(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    groups: Dict[Tuple[Any, ...], Dict[str, Any]] = {}

    for row in rows:
        if row.get("match_status") == "config_only":
            continue

        value, higher_is_better, metric_name = best_metric_value(row)
        if value is None:
            continue

        key = build_group_key(row)
        cur = groups.get(key)

        if cur is None:
            new_row = dict(row)
            new_row["_best_metric_name"] = metric_name
            new_row["_best_metric_value"] = value
            groups[key] = new_row
            continue

        cur_val = cur["_best_metric_value"]
        better = (value > cur_val) if higher_is_better else (value < cur_val)

        if better:
            new_row = dict(row)
            new_row["_best_metric_name"] = metric_name
            new_row["_best_metric_value"] = value
            groups[key] = new_row

    out = list(groups.values())
    out.sort(key=lambda x: (
        str(x.get("task_family")),
        str(x.get("exp_name")),
        str(x.get("method")),
        str(x.get("target_modules_tag")),
        to_int(x.get("rank")) if x.get("rank") is not None else 10**9,
    ))
    return out


# ============================================================
# Excel writing
# ============================================================

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9D9D9")


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=THIN_GRAY)
    ws.freeze_panes = "A2"


def autosize_columns(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for c in col_cells:
            v = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)


def add_table(ws):
    if ws.max_row < 2 or ws.max_column < 1:
        return
    end_col = get_column_letter(ws.max_column)
    ref = f"A1:{end_col}{ws.max_row}"
    name = re.sub(r"[^A-Za-z0-9_]", "_", f"{ws.title}_tbl")
    table = Table(displayName=name[:30], ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_sheet(ws, rows: List[Dict[str, Any]], columns: List[str]):
    ws.append(columns)
    for row in rows:
        ws.append([row.get(c) for c in columns])

    style_header(ws)
    autosize_columns(ws)
    add_table(ws)

    numeric_cols = {
        "rank", "alpha", "alpha_over_rank", "dropout",
        "seed", "num_labels",
        "batch_size", "max_length", "num_workers", "prefetch_factor", "pad_to_multiple_of",
        "lr", "weight_decay", "num_epochs", "max_train_steps", "grad_accum_steps", "warmup_ratio",
        "grad_clip", "log_every_steps", "eval_every_steps",
        "kd_T", "kd_lambda", "distill_temperature", "calibration_num_samples",

        "val_loss", "val_ppl", "val_accuracy", "val_acc", "val_f1", "val_macro_f1",
        "val_ce_loss", "val_kl_to_teacher", "val_mse_logits_to_teacher",
        "test_loss", "test_ppl", "test_accuracy", "test_acc", "test_f1", "test_macro_f1",
        "test_ce_loss", "test_kl_to_teacher", "test_mse_logits_to_teacher",
        "baseline_quant_test_loss", "baseline_quant_test_ppl",
        "baseline_quant_test_accuracy", "baseline_quant_test_acc", "baseline_quant_test_f1",
        "_best_metric_value",
    }

    header_map = {name: i + 1 for i, name in enumerate(columns)}
    for col_name, col_idx in header_map.items():
        if col_name in numeric_cols:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col_idx).number_format = "0.0000"


def add_readme_sheet(ws):
    rows = [
        ("Sheet", "Description"),
        ("merged_records", "Main table. Merged config + output records. Prefer this sheet."),
        ("configs_all", "All scanned config yaml files."),
        ("outputs_all", "All scanned output runs with metrics.json."),
        ("configs_missing_outputs", "Configs that were not matched to any output run."),
        ("outputs_missing_configs", "Output runs that were not matched to any config."),
        ("cls_records", "Subset of merged_records where task_family == cls."),
        ("lm_records", "Subset of merged_records where task_family == lm."),
        ("best_by_group", "Best matched runs grouped by task/exp/method/target_modules/teacher."),
        ("", ""),
        ("Key columns", "Meaning"),
        ("match_status", "matched / config_only / output_only"),
        ("method", "optimized / lora / eora / lora_kd / quantized / eora_quant / lora_quant / unknown"),
        ("output_dir_norm", "Normalized output_dir from config, relative to outputs_root."),
        ("outputs_rel_dir", "Actual output run directory relative to outputs_root."),
        ("target_modules_tag", "Parsed from path, e.g. tm-ap / tm-apf / tm-apfh."),
        ("teacher_alias", "Short teacher name derived from teacher_dir."),
        ("test_ppl", "Primary LM metric."),
        ("test_accuracy / test_f1", "Primary classification metrics."),
    ]
    for r in rows:
        ws.append(list(r))
    style_header(ws)
    autosize_columns(ws)


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--configs_root", type=str, default="configs", help="Configs root folder")
    parser.add_argument("--outputs_root", type=str, default="outputs", help="Outputs root folder")
    parser.add_argument("--out", type=str, default="outputs/experiment_summary.xlsx", help="Output xlsx path")
    parser.add_argument("--include_archive", action="store_true", help="Include configs/archive")
    args = parser.parse_args()

    configs_root = Path(args.configs_root).resolve()
    outputs_root = Path(args.outputs_root).resolve()
    out_path = Path(args.out).resolve()

    if not configs_root.exists():
        raise FileNotFoundError(f"configs_root not found: {configs_root}")
    if not outputs_root.exists():
        raise FileNotFoundError(f"outputs_root not found: {outputs_root}")

    # ----------------------------
    # Scan configs
    # ----------------------------
    config_files = sorted(configs_root.rglob("*.yaml"))
    config_rows: List[Dict[str, Any]] = []
    for p in config_files:
        if should_skip_config(p, configs_root, include_archive=args.include_archive):
            continue
        try:
            row = build_config_row(p, configs_root, outputs_root_name=outputs_root.name)
            config_rows.append(row)
        except Exception as e:
            print(f"[WARN] Failed to parse config {p}: {e}")

    # ----------------------------
    # Scan outputs
    # ----------------------------
    metrics_files = sorted(outputs_root.rglob("metrics.json"))
    output_rows: List[Dict[str, Any]] = []
    for p in metrics_files:
        try:
            row = build_output_row(p, outputs_root)
            output_rows.append(row)
        except Exception as e:
            print(f"[WARN] Failed to parse output {p}: {e}")

    # ----------------------------
    # Merge
    # ----------------------------
    merged_records, configs_missing_outputs, outputs_missing_configs = build_merged_records(config_rows, output_rows)

    merged_records.sort(key=lambda x: (
        str(x.get("task_family")),
        str(x.get("exp_name")),
        str(x.get("method")),
        str(x.get("match_status")),
        str(x.get("outputs_rel_dir") or x.get("config_rel_path")),
    ))

    cls_records = [r for r in merged_records if r.get("task_family") == "cls"]
    lm_records = [r for r in merged_records if r.get("task_family") == "lm"]
    best_rows = build_best_by_group(merged_records)

    # ----------------------------
    # Columns
    # ----------------------------
    preferred_columns = [
        "record_type", "match_status",

        "task_family", "exp_name", "method",
        "config_name", "run_name",

        "model_name", "task_type", "seed", "num_labels",

        "rank", "alpha", "alpha_over_rank", "dropout",
        "target_modules_tag", "target_modules",

        "output_dir", "output_dir_norm",
        "quant_output_dir", "quant_output_dir_norm",
        "export_best_model_dir", "export_best_model_dir_norm",

        "optimized_model_dir", "teacher_model_dir", "teacher_alias", "quantized_model_dir",

        "dataset_name",
        "train_split", "eval_split", "test_split",
        "train_local_disk_path", "eval_local_disk_path", "test_local_disk_path",
        "train_text_key", "eval_text_key", "test_text_key",

        "batch_size", "max_length", "num_workers", "pin_memory", "persistent_workers", "prefetch_factor", "pad_to_multiple_of",

        "lr", "weight_decay", "num_epochs", "max_train_steps", "grad_accum_steps", "warmup_ratio",
        "scheduler", "use_amp", "grad_clip", "log_every_steps", "eval_every_steps",

        "kd_T", "kd_lambda", "kd_loss",
        "distill_temperature",
        "calibration_num_samples",

        "val_loss", "val_ppl", "val_accuracy", "val_acc", "val_f1", "val_macro_f1",
        "val_ce_loss", "val_kl_to_teacher", "val_mse_logits_to_teacher",

        "test_loss", "test_ppl", "test_accuracy", "test_acc", "test_f1", "test_macro_f1",
        "test_ce_loss", "test_kl_to_teacher", "test_mse_logits_to_teacher",

        "baseline_quant_test_loss", "baseline_quant_test_ppl",
        "baseline_quant_test_accuracy", "baseline_quant_test_acc", "baseline_quant_test_f1",

        "outputs_rel_dir", "outputs_abs_dir",
        "config_rel_path", "config_abs_path",
        "metrics_json", "meta_json", "run_info_json", "config_used_json", "config_used_yaml",
    ]

    all_keys = set()
    for table_rows in [merged_records, config_rows, output_rows, best_rows]:
        for r in table_rows:
            all_keys.update(r.keys())
    extra_columns = [k for k in sorted(all_keys) if k not in preferred_columns]
    columns = preferred_columns + extra_columns

    # ----------------------------
    # Workbook
    # ----------------------------
    wb = Workbook()

    ws = wb.active
    ws.title = "merged_records"
    write_sheet(ws, merged_records, columns)

    ws = wb.create_sheet("configs_all")
    write_sheet(ws, config_rows, columns)

    ws = wb.create_sheet("outputs_all")
    write_sheet(ws, output_rows, columns)

    ws = wb.create_sheet("configs_missing_outputs")
    write_sheet(ws, configs_missing_outputs, columns)

    ws = wb.create_sheet("outputs_missing_configs")
    write_sheet(ws, outputs_missing_configs, columns)

    ws = wb.create_sheet("cls_records")
    write_sheet(ws, cls_records, columns)

    ws = wb.create_sheet("lm_records")
    write_sheet(ws, lm_records, columns)

    ws = wb.create_sheet("best_by_group")
    best_columns = ["_best_metric_name", "_best_metric_value"] + columns
    write_sheet(ws, best_rows, best_columns)

    ws = wb.create_sheet("readme")
    add_readme_sheet(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb.save(out_path)
    except PermissionError:
        print(f"[ERROR] Cannot write to: {out_path}")
        print("[ERROR] The Excel file is probably open in Excel/WPS.")
        print("[ERROR] Close it or choose a different --out path.")
        raise

    methods_found = sorted({str(r.get("method")) for r in merged_records})
    print(f"[Done] Config files scanned: {len(config_rows)}")
    print(f"[Done] metrics.json files scanned: {len(metrics_files)}")
    print(f"[Done] merged records: {len(merged_records)}")
    print(f"[Done] configs without outputs: {len(configs_missing_outputs)}")
    print(f"[Done] outputs without configs: {len(outputs_missing_configs)}")
    print(f"[Done] Methods found: {methods_found}")
    print(f"[Done] Saved Excel to: {out_path}")


if __name__ == "__main__":
    main()